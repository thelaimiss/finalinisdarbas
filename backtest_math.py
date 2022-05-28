from openpyxl import load_workbook

wb = load_workbook("etf.xlsx")

def how_many_line_has_excel():
    ws = wb["backtest"]
    rows=1
    checkeris = ws.cell(row=rows, column=1).value

    while checkeris != None:
        checkeris = ws.cell(row=rows, column=1).value
        if checkeris != None:
            rows+=1
    return rows


def simuliator():

    ws = wb["backtest"]
    etf_symbol_buye = ["empty", "empty"]
    left_no_used_money = [0, 0]
    change_bank = [0, 0]
    contract_size = [0, 0]
    data_now_list=[]

    check_of_data=[0]

    addonsbool=False


    addons=600

    row_number_of_backtest = 2



    def get_data_info():
        data=ws.cell(row=row_number_of_backtest, column=1).value

        characters_to_remove = "abcdefghijklmnopqrstuvwxyz ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        data_now = data
        for character in characters_to_remove:
            data_now = data_now.replace(character, "")
        if len(data_now_list)==0:
            data_now_list.append(data_now)
        if data_now_list[-1] != data_now:
            data_now_list.append(data_now)





    while True:
        rows = how_many_line_has_excel()


        if row_number_of_backtest >= rows:
            break

        get_data_info()

        check_of_data.append(data_now_list[-1])
        if len(check_of_data) >= 3:

            if check_of_data[-1] != check_of_data[-2]:
                addonsbool=True

            if addonsbool == True:

                ad_one=left_no_used_money[0] + addons
                ad_two = left_no_used_money[1] + addons
                left_no_used_money[0]=ad_one
                left_no_used_money[1]=ad_two

                addonsbool=False



        etf_trade_bank_first = ws["n1"].value
        etf_trade_bank_two = ws["n2"].value

        price = ws[f"e{row_number_of_backtest}"].value

        action = ws[f"d{row_number_of_backtest}"].value
        etf_symbol_file = ws[f"b{row_number_of_backtest}"].value

        # first etf data
        amount_kon_first = int(etf_trade_bank_first / price)
        used_money_first = round(price * amount_kon_first, 2)
        left_money_first = round(etf_trade_bank_first - used_money_first, 2)

        # two etf data

        amount_kon_two = int(etf_trade_bank_two / price)
        used_money_two = round(price * amount_kon_two, 2)
        left_money_two = round(etf_trade_bank_two - used_money_two, 2)

        #balansavimo ifas





        if action == "BUY" and etf_symbol_buye[0] == "empty":

            ws.cell(row=row_number_of_backtest, column=6).value = amount_kon_first
            ws.cell(row=row_number_of_backtest, column=7).value = used_money_first
            ws.cell(row=row_number_of_backtest, column=8).value = left_money_first
            ws.cell(row=row_number_of_backtest, column=10).value = "1 puse"

            left_no_used_money[0] = left_money_first
            etf_symbol_buye[0] = etf_symbol_file
            contract_size[0] = amount_kon_first

            row_number_of_backtest +=1
            continue

        if action == "BUY" and etf_symbol_buye[1] == "empty":
            ws.cell(row=row_number_of_backtest, column=6).value = amount_kon_two
            ws.cell(row=row_number_of_backtest, column=7).value = used_money_two
            ws.cell(row=row_number_of_backtest, column=8).value = left_money_two
            ws.cell(row=row_number_of_backtest, column=10).value = "2 puse"

            left_no_used_money[1] = left_money_two
            etf_symbol_buye[1] = etf_symbol_file
            contract_size[1] = amount_kon_two

            row_number_of_backtest +=1
            continue

        # viskas pirmam skyriui banko
        if action == "SELL" and etf_symbol_buye[0] == etf_symbol_file:
            # kintamasis skaiciavimui
            change_bank[0] = contract_size[0] * price + left_no_used_money[0] - 2
            # cashback-full suma po sell
            ws.cell(row=row_number_of_backtest, column=9).value = change_bank[0]
            # kuri puse banko
            ws.cell(row=row_number_of_backtest, column=10).value = "1 puse"
            # trade bank update

            ws.cell(row=1, column=14).value = change_bank[0]
            # emptinam savo laikoma data nes ivyko sell
            etf_symbol_buye[0] = "empty"
            left_no_used_money[0] = 0
            change_bank[0] = 0
            contract_size[0] = 0

            row_number_of_backtest +=1
            continue

            # viskas antram skyriui banko
        if action == "SELL" and etf_symbol_buye[1] == etf_symbol_file:
            # kintamasis skaiciavimui
            change_bank[1] = contract_size[1] * price + left_no_used_money[1] - 2
            # cashback-full suma po sell
            ws.cell(row=row_number_of_backtest, column=9).value = change_bank[1]
            # kuri puse banko
            ws.cell(row=row_number_of_backtest, column=10).value = "2 puse"

            # trade bank update

            ws.cell(row=2, column=14).value = change_bank[1]
            # emptinam savo laikoma data nes ivyko sell
            etf_symbol_buye[1] = "empty"
            left_no_used_money[1] = 0
            change_bank[1] = 0
            contract_size[1] = 0

            if etf_symbol_buye[0] == "empty" and etf_symbol_buye[1] == "empty":
                suma = etf_trade_bank_first + etf_trade_bank_two
                suma_finnaly = suma / 2

                ws.cell(row=2, column=14).value = suma_finnaly
                ws.cell(row=1, column=14).value = suma_finnaly

            row_number_of_backtest +=1
            continue

def data_bank_size_drow():

    ws = wb["backtest"]
    row=2
    while row <=100:

        chk_cshb_value = ws.cell(row=row, column=9).value
        bank_name = ws.cell(row=row, column=10).value


        if chk_cshb_value != None:
            if bank_name == "1 puse":
                ws.cell(row=row, column=30).value = chk_cshb_value
                row+=1

        if chk_cshb_value != None:
            if bank_name == "2 puse":
                ws.cell(row=row, column=31).value = chk_cshb_value
                row+=1

        row+=1



simuliator()
data_bank_size_drow()

wb.save("etf.xlsx")
