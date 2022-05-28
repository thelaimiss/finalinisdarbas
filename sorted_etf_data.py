from openpyxl import load_workbook

wb = load_workbook("etf.xlsx")

sheet_names = wb.sheetnames


def how_many_lines_has_excel():

    ws = wb[f"{sheet_names[0]}"]

    a = 1
    size = []

    while True:
        line = ws[f"f{a}"].value

        if line != None:
            a += 1
            size.append(a)
        else:
            break
    size_line = size[-1] - 1
    return size_line


a = how_many_lines_has_excel()


power_info = {
    f"{sheet_names[0]}": 0,  # dje 0
    f"{sheet_names[1]}": 0,  # eqqq 1
    f"{sheet_names[2]}": 0,  # iusa 2
    f"{sheet_names[3]}": 0,  # xsps 3
    f"{sheet_names[4]}": 0,  # ius3 4
    f"{sheet_names[5]}": 0,  # rtwo 5
    f"{sheet_names[6]}": 0,  # rtwo 5
    f"{sheet_names[7]}": 0,  # rtwo 5
    f"{sheet_names[8]}": 0,  # rtwo 5
    f"{sheet_names[9]}": 0,  # rtwo 5
}


close_price = {
    f"{sheet_names[0]}": 0,  # dje 0
    f"{sheet_names[1]}": 0,  # eqqq 1
    f"{sheet_names[2]}": 0,  # iusa 2
    f"{sheet_names[3]}": 0,  # xsps 3
    f"{sheet_names[4]}": 0,  # ius3 4
    f"{sheet_names[5]}": 0,  # rtwo 5
    f"{sheet_names[6]}": 0,  # rtwo 5
    f"{sheet_names[7]}": 0,  # rtwo 5
    f"{sheet_names[8]}": 0,  # rtwo 5
    f"{sheet_names[9]}": 0,  # rtwo 5
}

sorted_date = {
    f"{sheet_names[0]}": 0,  # dje 0
}


def zero_info():

    ws = wb[f"{sheet_names[0]}"]

    date = ws[f"a{a}"].value

    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({sheet_names[0]: power})
    close_price.update({sheet_names[0]: close})
    sorted_date.update({sheet_names[0]: date})


def one_info():

    ws = wb[f"{sheet_names[1]}"]

    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({sheet_names[1]: power})
    close_price.update({sheet_names[1]: close})


def two_info():

    ws = wb[f"{sheet_names[2]}"]

    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({sheet_names[2]: power})
    close_price.update({sheet_names[2]: close})


def three_info():

    ws = wb[f"{sheet_names[3]}"]

    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({sheet_names[3]: power})
    close_price.update({sheet_names[3]: close})


def four_info():

    ws = wb[f"{sheet_names[4]}"]

    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({sheet_names[4]: power})
    close_price.update({sheet_names[4]: close})


def five_info():
    ws = wb[f"{sheet_names[5]}"]

    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({sheet_names[5]: power})
    close_price.update({sheet_names[5]: close})


def six_info():
    ws = wb[f"{sheet_names[6]}"]

    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({sheet_names[6]: power})
    close_price.update({sheet_names[6]: close})


def seven_info():
    ws = wb[f"{sheet_names[7]}"]

    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({sheet_names[7]: power})
    close_price.update({sheet_names[7]: close})


def eight_info():
    ws = wb[f"{sheet_names[8]}"]

    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({sheet_names[8]: power})
    close_price.update({sheet_names[8]: close})


def nine_info():
    ws = wb[f"{sheet_names[9]}"]

    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({sheet_names[9]: power})
    close_price.update({sheet_names[9]: close})


def write_to_exel_etf_power_and_names():

    ws = wb[f"{sheet_names[10]}"]

    # date
    ws.cell(row=a, column=1).value = sorted_date[sheet_names[0]]

    # top 1 etf
    ws.cell(row=a, column=2).value = tp_one_convert[0]
    ws.cell(row=a, column=4).value = tp_one_convert[1]
    # top 2 etf
    ws.cell(row=a, column=5).value = tp_two_convert[0]
    ws.cell(row=a, column=7).value = tp_two_convert[1]
    # tp 3 etf
    ws.cell(row=a, column=8).value = tp_three_convert[0]
    ws.cell(row=a, column=10).value = tp_three_convert[1]

    # tp 4 etf
    ws.cell(row=a, column=12).value = tp_four_convert[0]
    ws.cell(row=a, column=14).value = tp_four_convert[1]

    # tp 5 etf
    ws.cell(row=a, column=15).value = tp_five_convert[0]
    ws.cell(row=a, column=17).value = tp_five_convert[1]

    # tp 6 etf
    ws.cell(row=a, column=18).value = tp_six_convert[0]
    ws.cell(row=a, column=20).value = tp_six_convert[1]
    # --------------------------------------------------
    # tp 7 etf
    ws.cell(row=a, column=21).value = tp_seven_convert[0]  # name
    ws.cell(row=a, column=23).value = tp_seven_convert[1]  # power

    # tp 8 etf
    ws.cell(row=a, column=24).value = tp_eight_convert[0]
    ws.cell(row=a, column=26).value = tp_eight_convert[1]

    # tp 9 etf
    ws.cell(row=a, column=27).value = tp_nine_convert[0]
    ws.cell(row=a, column=29).value = tp_nine_convert[1]

    # tp 10 etf
    ws.cell(row=a, column=30).value = tp_ten_convert[0]
    ws.cell(row=a, column=32).value = tp_ten_convert[1]


def price_writer():
    ws = wb[f"{sheet_names[10]}"]

    if tp_one_convert[0] == sheet_names[0]:  # eqq
        ws.cell(row=a, column=3).value = close_price[sheet_names[0]]

    if tp_one_convert[0] == sheet_names[1]:  # eqqq
        ws.cell(row=a, column=3).value = close_price[sheet_names[1]]

    if tp_one_convert[0] == sheet_names[2]:  # iusa
        ws.cell(row=a, column=3).value = close_price[sheet_names[2]]

    if tp_one_convert[0] == sheet_names[3]:  # exsps
        ws.cell(row=a, column=3).value = close_price[sheet_names[3]]

    if tp_one_convert[0] == sheet_names[4]:  # ius3
        ws.cell(row=a, column=3).value = close_price[sheet_names[4]]

    if tp_one_convert[0] == sheet_names[5]:  # rtwo
        ws.cell(row=a, column=3).value = close_price[sheet_names[5]]

    if tp_one_convert[0] == sheet_names[6]:  # rtwo
        ws.cell(row=a, column=3).value = close_price[sheet_names[6]]

    if tp_one_convert[0] == sheet_names[7]:  # rtwo
        ws.cell(row=a, column=3).value = close_price[sheet_names[7]]

    if tp_one_convert[0] == sheet_names[8]:  # rtwo
        ws.cell(row=a, column=3).value = close_price[sheet_names[8]]

    if tp_one_convert[0] == sheet_names[9]:  # rtwo
        ws.cell(row=a, column=3).value = close_price[sheet_names[9]]

    if tp_two_convert[0] == sheet_names[0]:
        ws.cell(row=a, column=6).value = close_price[sheet_names[0]]

    if tp_two_convert[0] == sheet_names[1]:
        ws.cell(row=a, column=6).value = close_price[sheet_names[1]]

    if tp_two_convert[0] == sheet_names[2]:
        ws.cell(row=a, column=6).value = close_price[sheet_names[2]]

    if tp_two_convert[0] == sheet_names[3]:
        ws.cell(row=a, column=6).value = close_price[sheet_names[3]]

    if tp_two_convert[0] == sheet_names[4]:
        ws.cell(row=a, column=6).value = close_price[sheet_names[4]]

    if tp_two_convert[0] == sheet_names[5]:
        ws.cell(row=a, column=6).value = close_price[sheet_names[5]]

    if tp_two_convert[0] == sheet_names[6]:
        ws.cell(row=a, column=6).value = close_price[sheet_names[6]]

    if tp_two_convert[0] == sheet_names[7]:
        ws.cell(row=a, column=6).value = close_price[sheet_names[7]]

    if tp_two_convert[0] == sheet_names[8]:
        ws.cell(row=a, column=6).value = close_price[sheet_names[8]]

    if tp_two_convert[0] == sheet_names[9]:
        ws.cell(row=a, column=6).value = close_price[sheet_names[9]]

    if tp_three_convert[0] == sheet_names[0]:
        ws.cell(row=a, column=9).value = close_price[sheet_names[0]]

    if tp_three_convert[0] == sheet_names[1]:
        ws.cell(row=a, column=9).value = close_price[sheet_names[1]]

    if tp_three_convert[0] == sheet_names[2]:
        ws.cell(row=a, column=9).value = close_price[sheet_names[2]]

    if tp_three_convert[0] == sheet_names[3]:
        ws.cell(row=a, column=9).value = close_price[sheet_names[3]]

    if tp_three_convert[0] == sheet_names[4]:
        ws.cell(row=a, column=9).value = close_price[sheet_names[4]]

    if tp_three_convert[0] == sheet_names[5]:
        ws.cell(row=a, column=9).value = close_price[sheet_names[5]]

    if tp_three_convert[0] == sheet_names[6]:
        ws.cell(row=a, column=9).value = close_price[sheet_names[6]]

    if tp_three_convert[0] == sheet_names[7]:
        ws.cell(row=a, column=9).value = close_price[sheet_names[7]]

    if tp_three_convert[0] == sheet_names[8]:
        ws.cell(row=a, column=9).value = close_price[sheet_names[8]]

    if tp_three_convert[0] == sheet_names[9]:
        ws.cell(row=a, column=9).value = close_price[sheet_names[9]]

    if tp_four_convert[0] == sheet_names[0]:
        ws.cell(row=a, column=13).value = close_price[sheet_names[0]]

    if tp_four_convert[0] == sheet_names[1]:
        ws.cell(row=a, column=13).value = close_price[sheet_names[1]]

    if tp_four_convert[0] == sheet_names[2]:
        ws.cell(row=a, column=13).value = close_price[sheet_names[2]]

    if tp_four_convert[0] == sheet_names[3]:
        ws.cell(row=a, column=13).value = close_price[sheet_names[3]]

    if tp_four_convert[0] == sheet_names[4]:
        ws.cell(row=a, column=13).value = close_price[sheet_names[4]]

    if tp_four_convert[0] == sheet_names[5]:
        ws.cell(row=a, column=13).value = close_price[sheet_names[5]]

    if tp_four_convert[0] == sheet_names[6]:
        ws.cell(row=a, column=13).value = close_price[sheet_names[6]]

    if tp_four_convert[0] == sheet_names[7]:
        ws.cell(row=a, column=13).value = close_price[sheet_names[7]]

    if tp_four_convert[0] == sheet_names[8]:
        ws.cell(row=a, column=13).value = close_price[sheet_names[8]]

    if tp_four_convert[0] == sheet_names[9]:
        ws.cell(row=a, column=13).value = close_price[sheet_names[9]]

    if tp_five_convert[0] == sheet_names[0]:
        ws.cell(row=a, column=16).value = close_price[sheet_names[0]]

    if tp_five_convert[0] == sheet_names[1]:
        ws.cell(row=a, column=16).value = close_price[sheet_names[1]]

    if tp_five_convert[0] == sheet_names[2]:
        ws.cell(row=a, column=16).value = close_price[sheet_names[2]]

    if tp_five_convert[0] == sheet_names[3]:
        ws.cell(row=a, column=16).value = close_price[sheet_names[3]]

    if tp_five_convert[0] == sheet_names[4]:
        ws.cell(row=a, column=16).value = close_price[sheet_names[4]]

    if tp_five_convert[0] == sheet_names[5]:
        ws.cell(row=a, column=16).value = close_price[sheet_names[5]]

    if tp_five_convert[0] == sheet_names[6]:
        ws.cell(row=a, column=16).value = close_price[sheet_names[6]]

    if tp_five_convert[0] == sheet_names[7]:
        ws.cell(row=a, column=16).value = close_price[sheet_names[7]]

    if tp_five_convert[0] == sheet_names[8]:
        ws.cell(row=a, column=16).value = close_price[sheet_names[8]]

    if tp_five_convert[0] == sheet_names[9]:
        ws.cell(row=a, column=16).value = close_price[sheet_names[9]]

    if tp_six_convert[0] == sheet_names[0]:
        ws.cell(row=a, column=19).value = close_price[sheet_names[0]]

    if tp_six_convert[0] == sheet_names[1]:
        ws.cell(row=a, column=19).value = close_price[sheet_names[1]]

    if tp_six_convert[0] == sheet_names[2]:
        ws.cell(row=a, column=19).value = close_price[sheet_names[2]]

    if tp_six_convert[0] == sheet_names[3]:
        ws.cell(row=a, column=19).value = close_price[sheet_names[3]]

    if tp_six_convert[0] == sheet_names[4]:
        ws.cell(row=a, column=19).value = close_price[sheet_names[4]]

    if tp_six_convert[0] == sheet_names[5]:
        ws.cell(row=a, column=19).value = close_price[sheet_names[5]]

    if tp_six_convert[0] == sheet_names[6]:
        ws.cell(row=a, column=19).value = close_price[sheet_names[6]]

    if tp_six_convert[0] == sheet_names[7]:
        ws.cell(row=a, column=19).value = close_price[sheet_names[7]]

    if tp_six_convert[0] == sheet_names[8]:
        ws.cell(row=a, column=19).value = close_price[sheet_names[8]]

    if tp_six_convert[0] == sheet_names[9]:
        ws.cell(row=a, column=19).value = close_price[sheet_names[9]]

    # --------------------------------------------

    if tp_seven_convert[0] == sheet_names[0]:
        ws.cell(row=a, column=22).value = close_price[sheet_names[0]]

    if tp_seven_convert[0] == sheet_names[1]:
        ws.cell(row=a, column=22).value = close_price[sheet_names[1]]

    if tp_seven_convert[0] == sheet_names[2]:
        ws.cell(row=a, column=22).value = close_price[sheet_names[2]]

    if tp_seven_convert[0] == sheet_names[3]:
        ws.cell(row=a, column=22).value = close_price[sheet_names[3]]

    if tp_seven_convert[0] == sheet_names[4]:
        ws.cell(row=a, column=22).value = close_price[sheet_names[4]]

    if tp_seven_convert[0] == sheet_names[5]:
        ws.cell(row=a, column=22).value = close_price[sheet_names[5]]

    if tp_seven_convert[0] == sheet_names[6]:
        ws.cell(row=a, column=22).value = close_price[sheet_names[6]]

    if tp_seven_convert[0] == sheet_names[7]:
        ws.cell(row=a, column=22).value = close_price[sheet_names[7]]

    if tp_seven_convert[0] == sheet_names[8]:
        ws.cell(row=a, column=22).value = close_price[sheet_names[8]]

    if tp_seven_convert[0] == sheet_names[9]:
        ws.cell(row=a, column=22).value = close_price[sheet_names[9]]

    if tp_eight_convert[0] == sheet_names[0]:
        ws.cell(row=a, column=25).value = close_price[sheet_names[0]]

    if tp_eight_convert[0] == sheet_names[1]:
        ws.cell(row=a, column=25).value = close_price[sheet_names[1]]

    if tp_eight_convert[0] == sheet_names[2]:
        ws.cell(row=a, column=25).value = close_price[sheet_names[2]]

    if tp_eight_convert[0] == sheet_names[3]:
        ws.cell(row=a, column=25).value = close_price[sheet_names[3]]

    if tp_eight_convert[0] == sheet_names[4]:
        ws.cell(row=a, column=25).value = close_price[sheet_names[4]]

    if tp_eight_convert[0] == sheet_names[5]:
        ws.cell(row=a, column=25).value = close_price[sheet_names[5]]

    if tp_eight_convert[0] == sheet_names[6]:
        ws.cell(row=a, column=25).value = close_price[sheet_names[6]]

    if tp_eight_convert[0] == sheet_names[7]:
        ws.cell(row=a, column=25).value = close_price[sheet_names[7]]

    if tp_eight_convert[0] == sheet_names[8]:
        ws.cell(row=a, column=25).value = close_price[sheet_names[8]]

    if tp_eight_convert[0] == sheet_names[9]:
        ws.cell(row=a, column=25).value = close_price[sheet_names[9]]

    if tp_nine_convert[0] == sheet_names[0]:
        ws.cell(row=a, column=28).value = close_price[sheet_names[0]]

    if tp_nine_convert[0] == sheet_names[1]:
        ws.cell(row=a, column=28).value = close_price[sheet_names[1]]

    if tp_nine_convert[0] == sheet_names[2]:
        ws.cell(row=a, column=28).value = close_price[sheet_names[2]]

    if tp_nine_convert[0] == sheet_names[3]:
        ws.cell(row=a, column=28).value = close_price[sheet_names[3]]

    if tp_nine_convert[0] == sheet_names[4]:
        ws.cell(row=a, column=28).value = close_price[sheet_names[4]]

    if tp_nine_convert[0] == sheet_names[5]:
        ws.cell(row=a, column=28).value = close_price[sheet_names[5]]

    if tp_nine_convert[0] == sheet_names[6]:
        ws.cell(row=a, column=28).value = close_price[sheet_names[6]]

    if tp_nine_convert[0] == sheet_names[7]:
        ws.cell(row=a, column=28).value = close_price[sheet_names[7]]

    if tp_nine_convert[0] == sheet_names[8]:
        ws.cell(row=a, column=28).value = close_price[sheet_names[8]]

    if tp_nine_convert[0] == sheet_names[9]:
        ws.cell(row=a, column=28).value = close_price[sheet_names[9]]

    if tp_ten_convert[0] == sheet_names[0]:
        ws.cell(row=a, column=31).value = close_price[sheet_names[0]]

    if tp_ten_convert[0] == sheet_names[1]:
        ws.cell(row=a, column=31).value = close_price[sheet_names[1]]

    if tp_ten_convert[0] == sheet_names[2]:
        ws.cell(row=a, column=31).value = close_price[sheet_names[2]]

    if tp_ten_convert[0] == sheet_names[3]:
        ws.cell(row=a, column=31).value = close_price[sheet_names[3]]

    if tp_ten_convert[0] == sheet_names[4]:
        ws.cell(row=a, column=31).value = close_price[sheet_names[4]]

    if tp_ten_convert[0] == sheet_names[5]:
        ws.cell(row=a, column=31).value = close_price[sheet_names[5]]

    if tp_ten_convert[0] == sheet_names[6]:
        ws.cell(row=a, column=31).value = close_price[sheet_names[6]]

    if tp_ten_convert[0] == sheet_names[7]:
        ws.cell(row=a, column=31).value = close_price[sheet_names[7]]

    if tp_ten_convert[0] == sheet_names[8]:
        ws.cell(row=a, column=31).value = close_price[sheet_names[8]]

    if tp_ten_convert[0] == sheet_names[9]:
        ws.cell(row=a, column=31).value = close_price[sheet_names[9]]


while True:

    if a >= 2:
        zero_info()
        one_info()
        two_info()
        three_info()
        four_info()
        five_info()
        six_info()
        seven_info()
        eight_info()
        nine_info()

        dict_sorted_values = sorted(
            power_info.items(), key=lambda kv: kv[1], reverse=True
        )
        print(dict_sorted_values)

        # only power

        top_one = dict_sorted_values[0]
        top_two = dict_sorted_values[1]
        top_three = dict_sorted_values[2]
        top_four = dict_sorted_values[3]
        top_five = dict_sorted_values[4]
        top_six = dict_sorted_values[5]

        top_seven = dict_sorted_values[6]
        top_eight = dict_sorted_values[7]
        top_nine = dict_sorted_values[8]
        top_ten = dict_sorted_values[9]

        tp_one = ",".join(map(str, top_one))
        tp_two = ",".join(map(str, top_two))
        tp_three = ",".join(map(str, top_three))
        tp_four = ",".join(map(str, top_four))
        tp_five = ",".join(map(str, top_five))
        tp_six = ",".join(map(str, top_six))

        tp_seven = ",".join(map(str, top_seven))
        tp_eight = ",".join(map(str, top_eight))
        tp_nine = ",".join(map(str, top_nine))
        tp_ten = ",".join(map(str, top_ten))

        tp_one_convert = tp_one.split(",")
        tp_two_convert = tp_two.split(",")
        tp_three_convert = tp_three.split(",")
        tp_four_convert = tp_four.split(",")
        tp_five_convert = tp_five.split(",")
        tp_six_convert = tp_six.split(",")

        tp_seven_convert = tp_seven.split(",")
        tp_eight_convert = tp_eight.split(",")
        tp_nine_convert = tp_nine.split(",")
        tp_ten_convert = tp_ten.split(",")

        write_to_exel_etf_power_and_names()
        price_writer()

        a -= 1
    else:
        print("duomenu pabaiga")
        break


wb.save("etf.xlsx")
