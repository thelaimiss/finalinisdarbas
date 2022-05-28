from openpyxl import load_workbook

wb = load_workbook("etf.xlsx")
#line of ETF file size
a = 614
sheet_names=wb.sheetnames
power_info = {
    f"{sheet_names[0]}": 0, #dje
    f"{sheet_names[1]}": 0,#eqqq
    f"{sheet_names[2]}": 0,#iusa
    f"{sheet_names[3]}": 0,#xsps
    f"{sheet_names[4]}": 0,#ius3
    f"{sheet_names[5]}": 0,#rtwo
}

open_price = {
    f"{sheet_names[0]}": 0, #dje
    f"{sheet_names[1]}": 0,#eqqq
    f"{sheet_names[2]}": 0,#iusa
    f"{sheet_names[3]}": 0,#xsps
    f"{sheet_names[4]}": 0,#ius3
    f"{sheet_names[5]}": 0,#rtwo
}

close_price = {
    f"{sheet_names[0]}": 0,  # dje
    f"{sheet_names[1]}": 0,  # eqqq
    f"{sheet_names[2]}": 0,  # iusa
    f"{sheet_names[3]}": 0,  # xsps
    f"{sheet_names[4]}": 0,  # ius3
    f"{sheet_names[5]}": 0,  # rtwo
}


def dje_info():
    ws = wb[f"{sheet_names[0]}"]

    # date = ws[f"a{a}"].value
    # open = ws[f"b{a}"].value
    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({"dje": power})
    open_price.update({"djeo": close})
    close_price.update({"djec": close})


def eqqq_info():
    ws = wb[f"{sheet_names[1]}"]

    # date = ws[f"a{a}"].value
    # open = ws[f"b{a}"].value
    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({"eqqq": power})
    open_price.update({"eqqqo": close})
    close_price.update({"eqqqc": close})


def iusa_info():
    ws = wb[f"{sheet_names[2]}"]

    # date = ws[f"a{a}"].value
    # open = ws[f"b{a}"].value
    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({"iusa": power})
    open_price.update({"iusao": close})
    close_price.update({"iusac": close})


def xsps_info():
    ws = wb[f"{sheet_names[3]}"]

    # date = ws[f"a{a}"].value
    # open = ws[f"b{a}"].value
    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({"xsps": power})
    open_price.update({"xspso": close})
    close_price.update({"xspsc": close})


def ius3_info():
    ws = wb[f"{sheet_names[4]}"]

    # date = ws[f"a{a}"].value
    # open = ws[f"b{a}"].value
    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({"ius3": power})
    open_price.update({"ius3o": close})
    close_price.update({"ius3c": close})


def rtwo_info():
    ws = wb[f"{sheet_names[5]}"]

    # date = ws[f"a{a}"].value
    # open = ws[f"b{a}"].value
    close = ws[f"c{a}"].value
    power = ws[f"f{a}"].value

    power_info.update({"rtwo": power})
    open_price.update({"rtwoo": close})
    close_price.update({"rtwoc": close})


def price_writer():
    ws = wb["simuliacija"]

    print(f"price {close_price['eqqqc']}," f"{top_one}")

    # only open price

    if tp_one_convert[0] == "dje":
        ws.cell(row=a, column=3).value = open_price["djeo"]

    if tp_one_convert[0] == "eqqq":
        ws.cell(row=a, column=3).value = open_price["eqqqo"]

    if tp_one_convert[0] == "iusa":
        ws.cell(row=a, column=3).value = open_price["iusao"]

    if tp_one_convert[0] == "xsps":
        ws.cell(row=a, column=3).value = open_price["xspso"]

    if tp_one_convert[0] == "ius3":
        ws.cell(row=a, column=3).value = open_price["ius3o"]

    if tp_one_convert[0] == "rtwo":
        ws.cell(row=a, column=3).value = open_price["rtwoo"]



    if tp_two_convert[0] == "dje":
        ws.cell(row=a, column=6).value = open_price["djeo"]

    if tp_two_convert[0] == "eqqq":
        ws.cell(row=a, column=6).value = open_price["eqqqo"]

    if tp_two_convert[0] == "iusa":
        ws.cell(row=a, column=6).value = open_price["iusao"]

    if tp_two_convert[0] == "xsps":
        ws.cell(row=a, column=6).value = open_price["xspso"]

    if tp_two_convert[0] == "ius3":
        ws.cell(row=a, column=6).value = open_price["ius3o"]

    if tp_two_convert[0] == "rtwo":
        ws.cell(row=a, column=6).value = open_price["rtwoo"]



    if tp_three_convert[0] == "dje":
        ws.cell(row=a, column=9).value = open_price["djeo"]

    if tp_three_convert[0] == "eqqq":
        ws.cell(row=a, column=9).value = open_price["eqqqo"]

    if tp_three_convert[0] == "iusa":
        ws.cell(row=a, column=9).value = open_price["iusao"]

    if tp_three_convert[0] == "xsps":
        ws.cell(row=a, column=9).value = open_price["xspso"]

    if tp_three_convert[0] == "ius3":
        ws.cell(row=a, column=9).value = open_price["ius3o"]

    if tp_three_convert[0] == "rtwo":
        ws.cell(row=a, column=9).value = open_price["rtwoo"]





    if tp_four_convert[0] == "dje":
        ws.cell(row=a, column=13).value = close_price["djec"]

    if tp_four_convert[0] == "eqqq":
        ws.cell(row=a, column=13).value = close_price["eqqqc"]

    if tp_four_convert[0] == "iusa":
        ws.cell(row=a, column=13).value = close_price["iusac"]

    if tp_four_convert[0] == "xsps":
        ws.cell(row=a, column=13).value = close_price["xspsc"]

    if tp_four_convert[0] == "ius3":
        ws.cell(row=a, column=13).value = close_price["ius3c"]

    if tp_four_convert[0] == "rtwo":
        ws.cell(row=a, column=13).value = close_price["rtwoc"]



    if tp_five_convert[0] == "dje":
        ws.cell(row=a, column=16).value = close_price["djec"]

    if tp_five_convert[0] == "eqqq":
        ws.cell(row=a, column=16).value = close_price["eqqqc"]

    if tp_five_convert[0] == "iusa":
        ws.cell(row=a, column=16).value = close_price["iusac"]

    if tp_five_convert[0] == "xsps":
        ws.cell(row=a, column=16).value = close_price["xspsc"]

    if tp_five_convert[0] == "ius3":
        ws.cell(row=a, column=16).value = close_price["ius3c"]

    if tp_five_convert[0] == "rtwo":
        ws.cell(row=a, column=16).value = close_price["rtwoc"]



    if tp_six_convert[0] == "dje":
        ws.cell(row=a, column=19).value = close_price["djec"]

    if tp_six_convert[0] == "eqqq":
        ws.cell(row=a, column=19).value = close_price["eqqqc"]

    if tp_six_convert[0] == "iusa":
        ws.cell(row=a, column=19).value = close_price["iusac"]

    if tp_six_convert[0] == "xsps":
        ws.cell(row=a, column=19).value = close_price["xspsc"]

    if tp_six_convert[0] == "ius3":
        ws.cell(row=a, column=19).value = close_price["ius3c"]

    if tp_six_convert[0] == "rtwo":
        ws.cell(row=a, column=19).value = close_price["rtwoc"]


def write_to_exel_etf_power():

    ws = wb["simuliacija"]
    # top 1 etf
    ws.cell(row=a, column=2).value = tp_one_convert[0]
    ws.cell(row=a, column=4).value = tp_one_convert[1]
    # top 2 etf
    ws.cell(row=a, column=5).value = tp_two_convert[0]
    ws.cell(row=a, column=7).value = tp_two_convert[1]
    # tp 3 etf
    ws.cell(row=a, column=8).value = tp_three_convert[0]
    ws.cell(row=a, column=10).value = tp_three_convert[1]

    # tp 4 etf
    ws.cell(row=a, column=12).value = tp_four_convert[0]
    ws.cell(row=a, column=14).value = tp_four_convert[1]

    # tp 5 etf
    ws.cell(row=a, column=15).value = tp_five_convert[0]
    ws.cell(row=a, column=17).value = tp_five_convert[1]

    # tp 6 etf
    ws.cell(row=a, column=18).value = tp_six_convert[0]
    ws.cell(row=a, column=20).value = tp_six_convert[1]


while True:

    if a >= 2:
        dje_info()
        eqqq_info()
        iusa_info()
        xsps_info()
        ius3_info()
        rtwo_info()

        dict_sorted_values = sorted(
            power_info.items(), key=lambda kv: kv[1], reverse=True
        )

        # only power

        top_one = dict_sorted_values[0]
        top_two = dict_sorted_values[1]
        top_three = dict_sorted_values[2]
        top_four = dict_sorted_values[3]
        top_five = dict_sorted_values[4]
        top_six = dict_sorted_values[5]

        tp_one = ",".join(map(str, top_one))
        tp_two = ",".join(map(str, top_two))
        tp_three = ",".join(map(str, top_three))
        tp_four = ",".join(map(str, top_four))
        tp_five = ",".join(map(str, top_five))
        tp_six = ",".join(map(str, top_six))

        tp_one_convert = tp_one.split(",")
        tp_two_convert = tp_two.split(",")
        tp_three_convert = tp_three.split(",")
        tp_four_convert = tp_four.split(",")
        tp_five_convert = tp_five.split(",")
        tp_six_convert = tp_six.split(",")

        print(
            f"TOP ONE:   {tp_one_convert}\n"
            f"TOP TWO:   {tp_two_convert}\n"
            f"TOP THREE: {tp_three_convert}\n"
            f"------------------------------"
        )

        write_to_exel_etf_power()
        price_writer()

        a -= 1
    else:
        print("duomenu pabaiga")
        break

wb.save("etf.xlsx")
