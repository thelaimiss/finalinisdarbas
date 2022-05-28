from openpyxl import load_workbook

wb = load_workbook("etf.xlsx")

def how_many_lines_has_excel():

    ws = wb["Reitingavimas"]

    a = 1
    size = []

    while True:
        line = ws[f"f{a}"].value

        if line != None:
            a += 1
            size.append(a)
        else:
            break
    size_line=size[-1]-1
    return  size_line


number_of_file_simuliator = how_many_lines_has_excel()


def data(a):
    ws = wb["Reitingavimas"]
    data = {
        "date_of_signal": ws[f"a{a}"].value,
        "top_one_etf": ws[f"b{a}"].value,
        "top_two_etf": ws[f"e{a}"].value,
        "top_three_etf": ws[f"h{a}"].value,
        "top_four_etf": ws[f"l{a}"].value,
        "top_five_etf": ws[f"o{a}"].value,
        "top_six_etf": ws[f"r{a}"].value,

        "top_seven_etf": ws[f"u{a}"].value,
        "top_eight_etf": ws[f"x{a}"].value,
        "top_nine_etf": ws[f"aa{a}"].value,
        "top_ten_etf": ws[f"ad{a}"].value,


        "top_one_price": ws[f"c{a}"].value,
        "top_two_price": ws[f"f{a}"].value,
        "top_three_price": ws[f"i{a}"].value,
        "top_four_price": ws[f"m{a}"].value,
        "top_five_price": ws[f"p{a}"].value,
        "top_six_price": ws[f"s{a}"].value,

        "top_seven_price": ws[f"v{a}"].value,
        "top_eight_price": ws[f"y{a}"].value,
        "top_nine_price": ws[f"ab{a}"].value,
        "top_ten_price": ws[f"ae{a}"].value,


        "top_one_etf_power": ws[f"d{a}"].value,
        "top_two_etf_power": ws[f"g{a}"].value,
        "top_three_etf_power": ws[f"j{a}"].value,
        "top_four_etf_power": ws[f"n{a}"].value,
        "top_five_etf_power": ws[f"q{a}"].value,
        "top_six_etf_power": ws[f"t{a}"].value,

        "top_seven_etf_power": ws[f"w{a}"].value,
        "top_eight_etf_power": ws[f"z{a}"].value,
        "top_nine_etf_power": ws[f"ac{a}"].value,
        "top_ten_etf_power": ws[f"af{a}"].value,


    }
    return data


def change_simulator_line():
    return number_of_file_simuliator


def get_dict_ifo(data):

    date_of_signal = data["date_of_signal"]

    top_one_etf = data["top_one_etf"]
    top_two_etf = data["top_two_etf"]
    top_three_etf = data["top_three_etf"]
    top_four_etf = data["top_four_etf"]
    top_five_etf = data["top_five_etf"]
    top_six_etf = data["top_six_etf"]

    top_seven_etf = data["top_seven_etf"]
    top_eight_etf = data["top_eight_etf"]
    top_nine_etf = data["top_nine_etf"]
    top_ten_etf = data["top_ten_etf"]

    top_one_price = data["top_one_price"]
    top_two_price = data["top_two_price"]
    top_three_price = data["top_three_price"]
    top_four_price = data["top_four_price"]
    top_five_price = data["top_five_price"]
    top_six_price = data["top_six_price"]

    top_seven_price = data["top_seven_price"]
    top_eight_price = data["top_eight_price"]
    top_nine_price = data["top_nine_price"]
    top_ten_price = data["top_ten_price"]

    top_one_etf_power = data["top_one_etf_power"]
    top_two_etf_power = data["top_two_etf_power"]
    top_three_etf_power = data["top_three_etf_power"]
    top_four_etf_power = data["top_four_etf_power"]
    top_five_etf_power = data["top_five_etf_power"]
    top_six_etf_power = data["top_six_etf_power"]

    top_seven_etf_power = data["top_seven_etf_power"]
    top_eight_etf_power = data["top_eight_etf_power"]
    top_nine_etf_power = data["top_nine_etf_power"]
    top_ten_etf_power = data["top_ten_etf_power"]

    return (
        top_one_etf,
        top_one_price,
        top_two_etf,
        top_two_price,
        top_three_etf,
        top_three_price,
        top_four_etf,
        top_four_price,
        top_five_etf,
        top_five_price,
        top_six_etf,
        top_six_price,
        top_seven_etf,
        top_seven_price,
        top_eight_etf,
        top_eight_price,
        top_nine_etf,
        top_nine_price,
        top_ten_etf,
        top_ten_price,

        top_one_etf_power,
        top_two_etf_power,
        top_three_etf_power,
        top_four_etf_power,
        top_five_etf_power,
        top_six_etf_power,
        top_seven_etf_power,
        top_eight_etf_power,
        top_nine_etf_power,
        top_ten_etf_power,

        date_of_signal,
    )


def the_brain(get_dict_ifo):

    multiple_trade_full = False

    global number_of_file_simuliator
    row_of_backtest = 2

    buy_list = ["first_etf", "another_etf"]

    while True:

        full_data_info = get_dict_ifo(data(change_simulator_line()))

        top_one_etf = full_data_info[0]
        top_one_price = full_data_info[1]

        top_two_etf = full_data_info[2]
        top_two_price = full_data_info[3]

        top_three_etf = full_data_info[4]
        top_three_price = full_data_info[5]

        top_four_etf = full_data_info[6]
        top_four_price = full_data_info[7]

        top_five_etf = full_data_info[8]
        top_five_price = full_data_info[9]

        top_six_etf = full_data_info[10]
        top_six_price = full_data_info[11]

        top_seven_etf = full_data_info[12]
        top_seven_price = full_data_info[13]

        top_eight_etf = full_data_info[14]
        top_eight_price = full_data_info[15]

        top_nine_etf = full_data_info[16]
        top_nine_price = full_data_info[17]

        top_ten_etf = full_data_info[18]
        top_ten_price = full_data_info[19]

        top_one_etf_power = full_data_info[20]
        top_two_etf_power = full_data_info[21]
        top_three_etf_power = full_data_info[22]
        top_four_etf_power = full_data_info[23]
        top_five_etf_power = full_data_info[24]
        top_six_etf_power = full_data_info[25]
        top_seven_etf_power = full_data_info[26]
        top_eight_etf_power = full_data_info[27]
        top_nine_etf_power = full_data_info[28]
        top_ten_etf_power = full_data_info[29]

        date_of_signal = full_data_info[30]

        if buy_list[0] == "empty" and buy_list[1] == "empty":
            multiple_trade_full = False

        if multiple_trade_full == False:
            ws = wb["backtest"]
            ws.cell(row=row_of_backtest, column=1).value = date_of_signal
            ws.cell(row=row_of_backtest, column=2).value = top_one_etf
            ws.cell(row=row_of_backtest, column=3).value = top_one_etf_power
            ws.cell(row=row_of_backtest, column=4).value = "BUY"
            ws.cell(row=row_of_backtest, column=5).value = top_one_price

            row_of_backtest += 1

            ws.cell(row=row_of_backtest, column=1).value = date_of_signal
            ws.cell(row=row_of_backtest, column=2).value = top_two_etf
            ws.cell(row=row_of_backtest, column=3).value = top_two_etf_power
            ws.cell(row=row_of_backtest, column=4).value = "BUY"
            ws.cell(row=row_of_backtest, column=5).value = top_two_price

            multiple_trade_full = True

            buy_list[0] = top_one_etf
            buy_list[1] = top_two_etf

            number_of_file_simuliator -= 1

            continue

        if buy_list[0] != top_one_etf or top_two_etf or top_three_etf:

            ws = wb["backtest"]

            if buy_list[0] == top_four_etf:

                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_four_etf
                ws.cell(row=row_of_backtest, column=3).value = top_four_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_four_price

                buy_list[0] = "empty"

            if buy_list[0] == top_five_etf:

                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_five_etf
                ws.cell(row=row_of_backtest, column=3).value = top_five_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_five_price

                buy_list[0] = "empty"

            if buy_list[0] == top_six_etf:

                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_six_etf
                ws.cell(row=row_of_backtest, column=3).value = top_six_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_six_price

                buy_list[0] = "empty"

            if buy_list[0] == top_seven_etf:
                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_seven_etf
                ws.cell(row=row_of_backtest, column=3).value = top_seven_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_seven_price

                buy_list[0] = "empty"

            if buy_list[0] == top_eight_etf:
                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_eight_etf
                ws.cell(row=row_of_backtest, column=3).value = top_eight_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_eight_price

                buy_list[0] = "empty"
            if buy_list[0] == top_nine_etf:
                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_nine_etf
                ws.cell(row=row_of_backtest, column=3).value = top_nine_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_nine_price

                buy_list[0] = "empty"

            if buy_list[0] == top_ten_etf:
                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_ten_etf
                ws.cell(row=row_of_backtest, column=3).value = top_ten_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_ten_price

                buy_list[0] = "empty"

        if buy_list[1] != top_one_etf or top_two_etf or top_three_etf:

            ws = wb["backtest"]

            if buy_list[1] == top_four_etf:

                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_four_etf
                ws.cell(row=row_of_backtest, column=3).value = top_four_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_four_price

                buy_list[1] = "empty"

            if buy_list[1] == top_five_etf:

                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_five_etf
                ws.cell(row=row_of_backtest, column=3).value = top_five_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_five_price

                buy_list[1] = "empty"

            if buy_list[1] == top_six_etf:

                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_six_etf
                ws.cell(row=row_of_backtest, column=3).value = top_six_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_six_price

                buy_list[1] = "empty"

            if buy_list[1] == top_seven_etf:

                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_seven_etf
                ws.cell(row=row_of_backtest, column=3).value = top_seven_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_seven_price

                buy_list[1] = "empty"
            if buy_list[1] == top_eight_etf:

                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_eight_etf
                ws.cell(row=row_of_backtest, column=3).value = top_eight_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_eight_price

                buy_list[1] = "empty"

            if buy_list[1] == top_nine_etf:

                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_nine_etf
                ws.cell(row=row_of_backtest, column=3).value = top_nine_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_nine_price

                buy_list[1] = "empty"

            if buy_list[1] == top_ten_etf:

                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_ten_etf
                ws.cell(row=row_of_backtest, column=3).value = top_ten_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "SELL"
                ws.cell(row=row_of_backtest, column=5).value = top_ten_price

                buy_list[1] = "empty"

            if buy_list[0] == "empty" and buy_list[1] != top_one_etf:

                ws = wb["backtest"]
                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_one_etf
                ws.cell(row=row_of_backtest, column=3).value = top_one_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "BUY"
                ws.cell(row=row_of_backtest, column=5).value = top_one_price

                buy_list[0] = top_one_etf

            if buy_list[0] == "empty" and buy_list[1] != top_two_etf:
                ws = wb["backtest"]
                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_two_etf
                ws.cell(row=row_of_backtest, column=3).value = top_two_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "BUY"
                ws.cell(row=row_of_backtest, column=5).value = top_two_price

                buy_list[0] = top_two_etf

            ################################################################

            if buy_list[1] == "empty" and buy_list[0] != top_one_etf:

                ws = wb["backtest"]
                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_one_etf
                ws.cell(row=row_of_backtest, column=3).value = top_one_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "BUY"
                ws.cell(row=row_of_backtest, column=5).value = top_one_price

                buy_list[1] = top_one_etf

            if buy_list[1] == "empty" and buy_list[0] != top_two_etf:
                ws = wb["backtest"]
                row_of_backtest += 1

                ws.cell(row=row_of_backtest, column=1).value = date_of_signal
                ws.cell(row=row_of_backtest, column=2).value = top_two_etf
                ws.cell(row=row_of_backtest, column=3).value = top_two_etf_power
                ws.cell(row=row_of_backtest, column=4).value = "BUY"
                ws.cell(row=row_of_backtest, column=5).value = top_two_price

                buy_list[1] = top_two_etf

            # stop while when end of line
            if number_of_file_simuliator <= 2:

                print("Job done. Now you can check the file.")
                break

        number_of_file_simuliator -= 1


the_brain(get_dict_ifo)
wb.save("etf.xlsx")
